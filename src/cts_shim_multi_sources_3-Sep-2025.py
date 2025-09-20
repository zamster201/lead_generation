import os
import re
import sqlite3
import argparse
from datetime import datetime
from typing import List, Dict, Any, Optional

import requests
import pandas as pd

# ---------------------------
# Regex & small helpers
# ---------------------------
URL_RE = re.compile(r"https?://[^\s)>\]\"']+", re.IGNORECASE)

def now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M")

def yyyymmdd(d: str) -> str:
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(d.replace("Z",""), fmt).strftime("%Y%m%d")
        except Exception:
            continue
    try:
        return pd.to_datetime(d).strftime("%Y%m%d")
    except Exception:
        return ""

def clean_src(src: str) -> str:
    return src.replace(".","").replace("_","").replace("/","").replace(" ","").upper()

def build_cts_id(source: str, posted_iso: str, id_: str) -> str:
    if not (source and posted_iso and id_):
        return ""
    stamp = yyyymmdd(posted_iso)
    if not stamp:
        return ""
    return f"CTS-{clean_src(source)}-{stamp}-{id_}".upper()

def extract_doc_urls_from_text(text: Optional[str]) -> List[str]:
    if not text or "http" not in text:
        return []
    urls = URL_RE.findall(text)
    out, seen = [], set()
    for u in urls:
        if u not in seen:
            seen.add(u); out.append(u)
    return out

# ---------------------------
# DB
# ---------------------------
def init_db(conn: sqlite3.Connection):
    cur = conn.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS leads (
        id TEXT PRIMARY KEY,
        cts_id TEXT,
        title TEXT,
        agency TEXT,
        posted TEXT,
        due TEXT,
        keywords TEXT,
        value_estimate TEXT,
        source TEXT,
        created TEXT,
        edited TEXT,
        status TEXT,
        product_match TEXT,
        feature TEXT,
        priority TEXT,
        partner_strategy TEXT,
        wf_status TEXT,
        active TEXT,
        next_action_date TEXT,
        owner TEXT,
        notes TEXT
    )""")
    cur.execute("""
    CREATE TABLE IF NOT EXISTS documents (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        lead_id TEXT,
        url TEXT,
        label TEXT,
        FOREIGN KEY(lead_id) REFERENCES leads(id)
    )""")
    conn.commit()

def upsert_lead(conn: sqlite3.Connection, row: Dict[str, Any]):
    cur = conn.cursor()
    cols = list(row.keys())
    placeholders = ",".join(["?"]*len(cols))
    update_clause = ",".join([f"{c}=excluded.{c}" for c in cols if c!="id"])
    sql = f"INSERT INTO leads ({','.join(cols)}) VALUES ({placeholders}) ON CONFLICT(id) DO UPDATE SET {update_clause}"
    cur.execute(sql, [row[c] for c in cols])

def insert_documents(conn: sqlite3.Connection, lead_id: str, urls: List[str]):
    if not urls: return
    cur = conn.cursor()
    for u in urls:
        cur.execute("INSERT INTO documents (lead_id, url, label) VALUES (?,?,?)", (lead_id, u, ""))

# ---------------------------
# SAM.gov API
# ---------------------------
import time

import time
import random

def fetch_sam(api_key: str, query: str, limit: int, posted_from: str, posted_to: str,
              max_retries: int = 5, throttle_ms: int = 1100, verbose: bool = False):
    if not api_key:
        print("WARN: SAM_API_KEY missing; skipping SAM.gov API pull.")
        return []

    def normalize_date(d: str) -> str:
        if not d: return ""
        try:
            if "-" in d:
                dt = datetime.strptime(d, "%Y-%m-%d")
            else:
                dt = datetime.strptime(d, "%m/%d/%Y")
            return dt.strftime("%m/%d/%Y")
        except Exception:
            return d

    url = "https://api.sam.gov/prod/opportunities/v2/search"
    params = {"api_key": api_key, "limit": limit}
    if query: params["q"] = query
    if posted_from: params["postedFrom"] = normalize_date(posted_from)
    if posted_to:   params["postedTo"]   = normalize_date(posted_to)

    attempt = 0
    base_sleep = 1.0      # seconds, starting backoff
    max_sleep  = 30.0     # cap per attempt
    steady_ms  = max(800, throttle_ms)  # min steady throttle ~0.8s

    while True:
        attempt += 1
        if verbose:
            print({"endpoint": url, "params": params, "attempt": attempt, "steady_ms": steady_ms})

        resp = requests.get(url, params=params, timeout=60)

        # 429: rate limit
        if resp.status_code == 429:
            ra = resp.headers.get("Retry-After")
            if ra:
                sleep_s = float(ra)
            else:
                # Exponential **full jitter**: sleep = rand(0, min(max_sleep, base_sleep * 2^attempt))
                sleep_s = random.uniform(0, min(max_sleep, base_sleep * (2 ** (attempt - 1))))
            if verbose:
                print(f"429 rate-limit. Sleeping {sleep_s:.2f}s (Retry-After={ra})")
            time.sleep(sleep_s)
            # Adapt steady-state throttle upward gently if we keep hitting 429s
            steady_ms = min(int(steady_ms * 1.25), 5000)  # cap at 5s
            continue

        # 5xx: transient server errors → backoff+jitter
        if 500 <= resp.status_code < 600:
            if attempt > max_retries:
                resp.raise_for_status()
            sleep_s = random.uniform(0, min(max_sleep, base_sleep * (2 ** (attempt - 1))))
            if verbose:
                print(f"{resp.status_code} server error. Sleeping {sleep_s:.2f}s")
            time.sleep(sleep_s)
            continue

        # Non-retryable or success
        resp.raise_for_status()
        data = resp.json().get("opportunitiesData", []) or []
        if verbose:
            rem = resp.headers.get("X-RateLimit-Remaining")
            print(f"SAM.gov returned {len(data)} items; remaining={rem}")

        # Gentle steady-state throttle after success
        time.sleep(steady_ms / 1000.0)
        return data


def normalize_sam_item(item: Dict[str, Any]) -> Dict[str, Any]:
    lead = {
        "id": item.get("noticeId") or item.get("solicitationNumber") or item.get("id") or "",
        "title": item.get("title") or "",
        "agency": item.get("agency") or item.get("department") or "",
        "posted": item.get("postedDate") or item.get("publishDate") or "",
        "due": item.get("responseDeadLine") or item.get("responseDate") or "",
        "keywords": ",".join(item.get("naicsCodes", []) or []),
        "value_estimate": item.get("baseType") or "Unknown",
        "source": "sam.gov",
        "created": now_ts(),
        "edited": now_ts(),
        "status": None,
        "product_match": None,
        "feature": None,
        "priority": None,
        "partner_strategy": None,
        "wf_status": "unfiled",
        "active": "Active",
        "next_action_date": None,
        "owner": None,
        "notes": None
    }
    lead["cts_id"] = build_cts_id(lead["source"], lead["posted"], lead["id"])
    return lead

def extract_doc_urls_sam(item: Dict[str, Any]) -> List[str]:
    urls = []
    for k in ("resourceLinks","links","attachments","documents","fileUrls","urls"):
        v = item.get(k)
        if isinstance(v, list):
            for e in v:
                if isinstance(e, str) and e.startswith("http"):
                    urls.append(e)
                elif isinstance(e, dict):
                    for kk in ("url","link","href","downloadUrl"):
                        u = e.get(kk)
                        if isinstance(u, str) and u.startswith("http"):
                            urls.append(u)
    for text_key in ("description","summary","text","body","solicitationDescription"):
        urls.extend(extract_doc_urls_from_text(item.get(text_key)))
    out, seen = [], set()
    for u in urls:
        if u not in seen:
            seen.add(u); out.append(u)
    return out

# ---------------------------
# File ingests (SEWP/NITAAC)
# ---------------------------
def read_table_file(path: str) -> pd.DataFrame:
    if not path:
        return pd.DataFrame()
    if not os.path.exists(path):
        print(f"WARN: File not found: {path}")
        return pd.DataFrame()
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return pd.read_excel(path)
    else:
        return pd.read_csv(path)

def normalize_sewp_row(row: Dict[str, Any]) -> Dict[str, Any]:
    id_     = row.get("Solicitation Number") or row.get("Solicitation") or row.get("RFQ") or row.get("ID") or ""
    title   = row.get("Title") or row.get("Requirement Title") or row.get("Description") or ""
    agency  = row.get("Agency") or row.get("Customer") or ""
    posted  = row.get("Release Date") or row.get("Posted") or row.get("Publish Date") or ""
    due     = row.get("Response Date") or row.get("Due Date") or ""
    value   = row.get("Ceiling") or row.get("Est. Value") or row.get("Value") or "Unknown"
    url     = row.get("URL") or row.get("Link") or row.get("Solicitation URL") or ""

    lead = {
        "id": str(id_).strip(),
        "title": str(title).strip(),
        "agency": str(agency).strip(),
        "posted": str(posted).strip(),
        "due": str(due).strip(),
        "keywords": "",
        "value_estimate": str(value).strip(),
        "source": "nasa_sewp",
        "created": now_ts(),
        "edited": now_ts(),
        "status": None,
        "product_match": None,
        "feature": None,
        "priority": None,
        "partner_strategy": None,
        "wf_status": "unfiled",
        "active": "Active",
        "next_action_date": None,
        "owner": None,
        "notes": None
    }
    lead["cts_id"] = build_cts_id(lead["source"], lead["posted"], lead["id"])
    doc_urls = []
    if url: doc_urls.append(str(url).strip())
    for k in ("Description","Long Description","Additional Info"):
        doc_urls += extract_doc_urls_from_text(str(row.get(k) or ""))
    deduped, seen = [], set()
    for u in doc_urls:
        if u not in seen:
            seen.add(u); deduped.append(u)
    return lead, deduped

def ingest_sewp_file(conn: sqlite3.Connection, path: str):
    df = read_table_file(path)
    if df.empty:
        return
    for _, s in df.iterrows():
        lead, urls = normalize_sewp_row(s.to_dict())
        if not lead["id"]:
            continue
        upsert_lead(conn, lead)
        insert_documents(conn, lead["id"], urls)

def normalize_nitaac_row(row: Dict[str, Any]) -> Dict[str, Any]:
    id_     = row.get("Solicitation Number") or row.get("RFQ") or row.get("ID") or ""
    title   = row.get("Title") or row.get("Requirement Title") or row.get("Description") or ""
    agency  = row.get("Agency") or row.get("Customer") or ""
    posted  = row.get("Posted") or row.get("Publish Date") or row.get("Release Date") or ""
    due     = row.get("Due Date") or row.get("Response Date") or ""
    value   = row.get("Ceiling") or row.get("Est. Value") or row.get("Value") or "Unknown"
    url     = row.get("URL") or row.get("Link") or row.get("Solicitation URL") or ""

    lead = {
        "id": str(id_).strip(),
        "title": str(title).strip(),
        "agency": str(agency).strip(),
        "posted": str(posted).strip(),
        "due": str(due).strip(),
        "keywords": "",
        "value_estimate": str(value).strip(),
        "source": "nitaac",
        "created": now_ts(),
        "edited": now_ts(),
        "status": None,
        "product_match": None,
        "feature": None,
        "priority": None,
        "partner_strategy": None,
        "wf_status": "unfiled",
        "active": "Active",
        "next_action_date": None,
        "owner": None,
        "notes": None
    }
    lead["cts_id"] = build_cts_id(lead["source"], lead["posted"], lead["id"])
    doc_urls = []
    if url: doc_urls.append(str(url).strip())
    for k in ("Description","Long Description","Additional Info"):
        doc_urls += extract_doc_urls_from_text(str(row.get(k) or ""))
    deduped, seen = [], set()
    for u in doc_urls:
        if u not in seen:
            seen.add(u); deduped.append(u)
    return lead, deduped

def ingest_nitaac_file(conn: sqlite3.Connection, path: str):
    df = read_table_file(path)
    if df.empty:
        return
    for _, s in df.iterrows():
        lead, urls = normalize_nitaac_row(s.to_dict())
        if not lead["id"]:
            continue
        upsert_lead(conn, lead)
        insert_documents(conn, lead["id"], urls)

# ---------------------------
# Exports
# ---------------------------
def export_csv(conn: sqlite3.Connection, out_path: str):
    # Leads with newline-separated URLs (char(10) = LF)
    leads_df = pd.read_sql_query("""
        SELECT l.*,
               GROUP_CONCAT(d.url, char(10)) AS doc_urls
        FROM leads l
        LEFT JOIN documents d ON l.id = d.lead_id
        GROUP BY l.id
        ORDER BY l.posted DESC
    """, conn)
    # One row per document
    docs_df = pd.read_sql_query("""
        SELECT l.id AS lead_id,
               l.cts_id,
               l.title,
               d.url
        FROM documents d
        JOIN leads l ON l.id = d.lead_id
        ORDER BY l.posted DESC, d.id
    """, conn)

    # Main leads CSV
    leads_df.to_csv(out_path, index=False, encoding="utf-8")

    # Companion documents CSV (same folder, _documents suffix)
    import os
    root, ext = os.path.splitext(out_path)
    docs_out = f"{root}_documents{ext}"
    docs_df.to_csv(docs_out, index=False, encoding="utf-8")

def export_xlsx(conn: sqlite3.Connection, out_path: str):
    from openpyxl.styles import Alignment

    # Leads with newline-separated URLs
    leads_df = pd.read_sql_query("""
        SELECT l.*,
               GROUP_CONCAT(d.url, char(10)) AS doc_urls
        FROM leads l
        LEFT JOIN documents d ON l.id = d.lead_id
        GROUP BY l.id
        ORDER BY l.posted DESC
    """, conn)
    # One row per document
    docs_df = pd.read_sql_query("""
        SELECT l.id AS lead_id,
            l.cts_id,
            l.title,
            d.url
        FROM documents d
        JOIN leads l ON l.id = d.lead_id
        ORDER BY l.posted DESC, d.id
    """, conn)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        leads_df.to_excel(writer, index=False, sheet_name="Leads")
        docs_df.to_excel(writer, index=False, sheet_name="Documents")

        wb = writer.book
        ws_leads = wb["Leads"]
        ws_docs = wb["Documents"]

        # Wrap text in the Leads!doc_urls column so LF shows as multi-line
        if "doc_urls" in leads_df.columns:
            doc_col_idx = list(leads_df.columns).index("doc_urls") + 1  # 1-based
            for r in range(2, ws_leads.max_row + 1):
                ws_leads.cell(r, doc_col_idx).alignment = Alignment(wrap_text=True)

        # Make Documents!url cells clickable hyperlinks
        # Assume last column is 'url'
        url_col_idx = docs_df.columns.get_loc("url") + 1
        for r in range(2, ws_docs.max_row + 1):
            c = ws_docs.cell(r, url_col_idx)
            c.hyperlink = c.value
            c.style = "Hyperlink"

def export_md(conn: sqlite3.Connection, out_dir: str):
    import os
    os.makedirs(out_dir, exist_ok=True)

    # Leads (no concat needed here; we’ll render bullets from docs)
    leads = pd.read_sql_query("""
        SELECT *
        FROM leads
        ORDER BY posted DESC
    """, conn)

    # All docs → group by lead_id
    docs = pd.read_sql_query("""
        SELECT lead_id, url
        FROM documents
        ORDER BY id
    """, conn)
    doc_map = {}
    for lead_id, url in docs.itertuples(index=False):
        doc_map.setdefault(lead_id, []).append(url)

    for _, r in leads.iterrows():
        created = r["created"] or now_ts()
        edited  = r["edited"] or now_ts()
        source  = r["source"] or "sam.gov"
        wf      = r["wf_status"] or "unfiled"
        tags    = f'["lead","{source}"]'
        title   = r["title"] or "Untitled"
        urls    = doc_map.get(r["id"], [])

        frontmatter = f"""created: {created}
edited: {edited}
origin: {source}
source: {source}
wf_status: {wf}
tags: {tags}
"""
        if urls:
            docs_md = "\n".join(f"- <{u}>" for u in urls)
        else:
            docs_md = "(none detected)"

        body = f"""# {title}

**ID**: {r['id']}
**CTS ID**: {r.get('cts_id','')}
**Agency**: {r.get('agency','')}
**Posted**: {r.get('posted','')}
**Due**: {r.get('due','')}
**Source**: {source}
**Value Estimate**: {r.get('value_estimate','')}
**Active**: {r.get('active','')}

---

### Keywords
{r.get('keywords','')}

### Documents / Links
{docs_md}

### Notes
{r.get('notes','') or ''}
"""
        fname = f"{(r['cts_id'] or r['id']).replace('/','-')}.md"
        with open(os.path.join(out_dir, fname), "w", encoding="utf-8") as f:
            f.write(frontmatter + "\n---\n\n" + body)

# ---------------------------
# Argparse (flags > env > defaults)
# ---------------------------
def parse_args():
    p = argparse.ArgumentParser(description="CTS ingest shim: SAM.gov API + SEWP/NITAAC files → SQLite → Exports")
    p.add_argument("--sam-api-key", default=os.getenv("SAM_API_KEY", ""), help="SAM.gov API key (overrides env)")
    p.add_argument("--query", default=os.getenv("CTS_QUERY", "cybersecurity"), help="SAM.gov query string")
    p.add_argument("--from", dest="posted_from", default=os.getenv("CTS_POSTED_FROM", ""), help="Posted from (YYYY-MM-DD)")
    p.add_argument("--to", dest="posted_to", default=os.getenv("CTS_POSTED_TO", ""), help="Posted to (YYYY-MM-DD)")
    p.add_argument("--limit", type=int, default=int(os.getenv("CTS_LIMIT", "100")), help="Max SAM results")
    p.add_argument("--sewp-file", default=os.getenv("CTS_SEWP_FILE", ""), help="Path to SEWP CSV/XLSX export")
    p.add_argument("--nitaac-file", default=os.getenv("CTS_NITAAC_FILE", ""), help="Path to NITAAC CSV/XLSX export")
    p.add_argument("--db-path", default=os.getenv("CTS_DB_PATH", "leads.db"), help="SQLite DB path")
    p.add_argument("--export-dir", default=os.getenv("CTS_EXPORT_DIR", "exports"), help="Exports folder")
    p.add_argument("--no-sam", action="store_true", help="Skip SAM.gov pull even if key is present")

 # 🔥 add these:
    p.add_argument("--max-retries", type=int, default=5, help="Max retries on 429/5xx")
    p.add_argument("--throttle-ms", type=int, default=1100, help="Sleep between calls (ms)")
    p.add_argument("--verbose", action="store_true", help="Log request params and counts")
    return p.parse_args()

# ---------------------------
# Main
# ---------------------------
def main():
    args = parse_args()

    api_key     = args.sam_api_key
    query       = args.query
    posted_from = args.posted_from
    posted_to   = args.posted_to
    limit       = args.limit
    sewp_file   = args.sewp_file
    nitaac_file = args.nitaac_file
    db_path     = args.db_path
    export_dir  = args.export_dir
    skip_sam    = args.no_sam

    os.makedirs(export_dir, exist_ok=True)
    conn = sqlite3.connect(db_path)
    init_db(conn)

    # 1) SAM.gov
    if not skip_sam:
        try:
            items = fetch_sam(
                api_key, query, limit, posted_from, posted_to,
                max_retries=args.max_retries,
                throttle_ms=args.throttle_ms,
                verbose=args.verbose
            )
            for it in items:
                lead = normalize_sam_item(it)
                if not lead["id"]:
                    continue
                upsert_lead(conn, lead)
                urls = extract_doc_urls_sam(it)
                insert_documents(conn, lead["id"], urls)
        except Exception as e:
            print(f"WARN: SAM.gov fetch failed: {e}")

    # 2) SEWP (file)
    if sewp_file:
        try:
            ingest_sewp_file(conn, sewp_file)
        except Exception as e:
            print(f"WARN: SEWP ingest failed: {e}")

    # 3) NITAAC (file)
    if nitaac_file:
        try:
            ingest_nitaac_file(conn, nitaac_file)
        except Exception as e:
            print(f"WARN: NITAAC ingest failed: {e}")

    conn.commit()

    stamp = datetime.now().strftime("%d-%b-%y")
    export_csv(conn,  os.path.join(export_dir, f"leads_{stamp}.csv"))
    export_xlsx(conn, os.path.join(export_dir, f"leads_{stamp}.xlsx"))
    export_md(conn,   os.path.join(export_dir, f"md_{stamp}"))
    conn.close()
    print(f"Done. DB: {db_path} | Exports in: {export_dir}")

if __name__ == "__main__":
    main()
